"""
Export Service

Handles conversion of results to various export formats:
- JSON (copy-paste, API integration)
- CSV/TSV (spreadsheets, analysis)
- Markdown (documentation, GitHub)
- PDF (reports, clinical)
- VCF (bioinformatics pipelines)
- Excel (enterprise)
"""

import json
import io
from datetime import datetime
from typing import Any

from ..models import (
    ExportFormat,
    GeneScore,
    PredictionResult,
    format_as_markdown,
    format_as_csv,
    format_as_tsv,
)


class ExportService:
    """Service for exporting results in various formats."""

    def export(
        self,
        data: dict[str, Any],
        format: ExportFormat,
        filename: str = "alphagenome_result",
    ) -> tuple[bytes, str, str]:
        """
        Export data in the specified format.

        Args:
            data: The data to export
            format: Target export format
            filename: Base filename (without extension)

        Returns:
            Tuple of (content_bytes, content_type, filename_with_extension)
        """
        if format == ExportFormat.JSON:
            return self._export_json(data, filename)
        elif format == ExportFormat.CSV:
            return self._export_csv(data, filename)
        elif format == ExportFormat.TSV:
            return self._export_tsv(data, filename)
        elif format == ExportFormat.MARKDOWN:
            return self._export_markdown(data, filename)
        elif format == ExportFormat.PDF:
            return self._export_pdf(data, filename)
        elif format == ExportFormat.VCF:
            return self._export_vcf(data, filename)
        elif format == ExportFormat.EXCEL:
            return self._export_excel(data, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_json(
        self, data: dict, filename: str
    ) -> tuple[bytes, str, str]:
        """Export as formatted JSON."""
        # Add export metadata
        export_data = {
            "export_metadata": {
                "format": "json",
                "generated_at": datetime.utcnow().isoformat(),
                "tool": "AlphaGenome Explorer",
                "version": "1.0.0",
            },
            **data,
        }

        content = json.dumps(export_data, indent=2, default=str)
        return (
            content.encode("utf-8"),
            "application/json",
            f"{filename}.json",
        )

    def _export_csv(
        self, data: dict, filename: str
    ) -> tuple[bytes, str, str]:
        """Export as CSV."""
        lines = []

        # Header comment
        lines.append(f"# AlphaGenome Explorer Export")
        lines.append(f"# Generated: {datetime.utcnow().isoformat()}")
        lines.append("")

        # If we have scores, export them
        if "scores" in data and data["scores"]:
            scores = data["scores"]
            if isinstance(scores[0], dict):
                # Get headers from first item
                headers = list(scores[0].keys())
                lines.append(",".join(headers))

                for score in scores:
                    values = [str(score.get(h, "")) for h in headers]
                    lines.append(",".join(values))
            else:
                # Assume GeneScore objects
                lines.append(
                    "gene_id,gene_name,strand,tissue,raw_score,quantile_score,interpretation"
                )
                for s in scores:
                    lines.append(
                        f"{s.gene_id},{s.gene_name},{s.strand},{s.tissue},"
                        f"{s.raw_score},{s.quantile_score},{s.interpretation}"
                    )

        # If we have summary
        elif "summary" in data:
            summary = data["summary"]
            lines.append("field,value")
            for key, value in summary.items():
                if isinstance(value, list):
                    value = ";".join(str(v) for v in value)
                lines.append(f"{key},{value}")

        content = "\n".join(lines)
        return (
            content.encode("utf-8"),
            "text/csv",
            f"{filename}.csv",
        )

    def _export_tsv(
        self, data: dict, filename: str
    ) -> tuple[bytes, str, str]:
        """Export as TSV."""
        lines = []

        lines.append(f"# AlphaGenome Explorer Export")
        lines.append(f"# Generated: {datetime.utcnow().isoformat()}")
        lines.append("")

        if "scores" in data and data["scores"]:
            scores = data["scores"]
            if isinstance(scores[0], dict):
                headers = list(scores[0].keys())
                lines.append("\t".join(headers))

                for score in scores:
                    values = [str(score.get(h, "")) for h in headers]
                    lines.append("\t".join(values))

        content = "\n".join(lines)
        return (
            content.encode("utf-8"),
            "text/tab-separated-values",
            f"{filename}.tsv",
        )

    def _export_markdown(
        self, data: dict, filename: str
    ) -> tuple[bytes, str, str]:
        """Export as Markdown."""
        lines = []

        # Header
        variant = data.get("request_params", {}).get("variant", "Analysis")
        lines.append(f"# AlphaGenome Analysis: {variant}")
        lines.append("")
        lines.append(f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
        lines.append("")

        # Summary
        if "summary" in data:
            summary = data["summary"]
            lines.append("## Summary")
            lines.append("")
            lines.append("| Metric | Value |")
            lines.append("|--------|-------|")

            if isinstance(summary, dict):
                for key, value in summary.items():
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value)
                    lines.append(f"| {key} | {value} |")

            lines.append("")

        # Scores table
        if "scores" in data and data["scores"]:
            lines.append("## Gene Scores")
            lines.append("")
            lines.append("| Gene | Tissue | Raw Score | Quantile | Interpretation |")
            lines.append("|------|--------|-----------|----------|----------------|")

            scores = data["scores"][:20]  # Top 20
            for s in scores:
                if isinstance(s, dict):
                    lines.append(
                        f"| {s.get('gene_name', 'N/A')} | {s.get('tissue', 'N/A')} | "
                        f"{s.get('raw_score', 0):.4f} | {s.get('quantile_score', 0):.2f} | "
                        f"{s.get('interpretation', 'N/A')} |"
                    )

            lines.append("")

        # Footer
        lines.append("---")
        lines.append("")
        lines.append("*Generated by [AlphaGenome Explorer](https://github.com/)*")
        lines.append("")
        lines.append("**How to cite:**")
        lines.append("> Avsec et al. \"Advancing regulatory variant effect prediction with AlphaGenome\" Nature 2026")

        content = "\n".join(lines)
        return (
            content.encode("utf-8"),
            "text/markdown",
            f"{filename}.md",
        )

    def _export_pdf(
        self, data: dict, filename: str
    ) -> tuple[bytes, str, str]:
        """Export as PDF report."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
                Table,
                TableStyle,
            )

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                "Title",
                parent=styles["Heading1"],
                fontSize=18,
                spaceAfter=20,
            )
            variant = data.get("request_params", {}).get("variant", "Analysis")
            story.append(Paragraph(f"AlphaGenome Analysis Report", title_style))
            story.append(Paragraph(f"Variant: {variant}", styles["Heading2"]))
            story.append(Spacer(1, 12))

            # Metadata
            story.append(
                Paragraph(
                    f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
                    styles["Normal"],
                )
            )
            story.append(Spacer(1, 20))

            # Summary section
            if "summary" in data:
                story.append(Paragraph("Summary", styles["Heading2"]))
                summary = data["summary"]
                if isinstance(summary, dict):
                    summary_data = [["Metric", "Value"]]
                    for key, value in summary.items():
                        if isinstance(value, list):
                            value = ", ".join(str(v) for v in value)
                        summary_data.append([key.replace("_", " ").title(), str(value)])

                    table = Table(summary_data, colWidths=[2 * inch, 4 * inch])
                    table.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, 0), 12),
                                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            ]
                        )
                    )
                    story.append(table)
                    story.append(Spacer(1, 20))

            # Scores section
            if "scores" in data and data["scores"]:
                story.append(Paragraph("Gene Scores (Top 15)", styles["Heading2"]))

                scores_data = [["Gene", "Tissue", "Score", "Quantile", "Interpretation"]]
                for s in data["scores"][:15]:
                    if isinstance(s, dict):
                        scores_data.append(
                            [
                                s.get("gene_name", "N/A"),
                                s.get("tissue", "N/A")[:20],
                                f"{s.get('raw_score', 0):.4f}",
                                f"{s.get('quantile_score', 0):.2f}",
                                s.get("interpretation", "N/A")[:25],
                            ]
                        )

                table = Table(
                    scores_data,
                    colWidths=[1.2 * inch, 1.5 * inch, 0.8 * inch, 0.8 * inch, 1.7 * inch],
                )
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 9),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ]
                    )
                )
                story.append(table)
                story.append(Spacer(1, 20))

            # Footer
            story.append(Spacer(1, 30))
            story.append(
                Paragraph(
                    "Generated by AlphaGenome Explorer | "
                    "Powered by Google DeepMind AlphaGenome API",
                    styles["Normal"],
                )
            )
            story.append(Spacer(1, 10))
            story.append(
                Paragraph(
                    "<i>Citation: Avsec et al. \"Advancing regulatory variant effect "
                    "prediction with AlphaGenome\" Nature 2026</i>",
                    styles["Normal"],
                )
            )

            doc.build(story)
            buffer.seek(0)

            return (
                buffer.read(),
                "application/pdf",
                f"{filename}.pdf",
            )

        except ImportError:
            # Fallback to text if reportlab not available
            content = f"PDF generation requires reportlab. Install with: pip install reportlab\n\n"
            content += self._export_markdown(data, filename)[0].decode("utf-8")
            return (
                content.encode("utf-8"),
                "text/plain",
                f"{filename}.txt",
            )

    def _export_vcf(
        self, data: dict, filename: str
    ) -> tuple[bytes, str, str]:
        """Export as annotated VCF."""
        lines = []

        # VCF header
        lines.append("##fileformat=VCFv4.2")
        lines.append(f"##fileDate={datetime.utcnow().strftime('%Y%m%d')}")
        lines.append("##source=AlphaGenomeExplorer")
        lines.append('##INFO=<ID=AG_IMPACT,Number=1,Type=String,Description="AlphaGenome impact prediction">')
        lines.append('##INFO=<ID=AG_GENE,Number=.,Type=String,Description="Affected genes">')
        lines.append('##INFO=<ID=AG_SCORE,Number=1,Type=Float,Description="AlphaGenome raw score">')
        lines.append('##INFO=<ID=AG_QUANTILE,Number=1,Type=Float,Description="AlphaGenome quantile score">')
        lines.append("#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO")

        # Parse variant from request params
        variant_str = data.get("request_params", {}).get("variant", "")
        if variant_str:
            parts = variant_str.split(":")
            if len(parts) >= 3:
                chrom = parts[0].replace("chr", "")
                pos = parts[1]
                ref_alt = parts[2].split(">")
                ref = ref_alt[0]
                alt = ref_alt[1] if len(ref_alt) > 1 else "."

                # Build INFO field
                info_parts = []
                summary = data.get("summary", {})
                if isinstance(summary, dict):
                    info_parts.append(f"AG_IMPACT={summary.get('impact_level', 'UNKNOWN')}")
                    genes = summary.get("affected_genes", [])
                    if genes:
                        info_parts.append(f"AG_GENE={','.join(genes)}")

                scores = data.get("scores", [])
                if scores:
                    top_score = scores[0] if isinstance(scores[0], dict) else {}
                    info_parts.append(f"AG_SCORE={top_score.get('raw_score', 0):.4f}")
                    info_parts.append(f"AG_QUANTILE={top_score.get('quantile_score', 0):.2f}")

                info = ";".join(info_parts) if info_parts else "."
                lines.append(f"{chrom}\t{pos}\t.\t{ref}\t{alt}\t.\tPASS\t{info}")

        content = "\n".join(lines)
        return (
            content.encode("utf-8"),
            "text/plain",
            f"{filename}.vcf",
        )

    def _export_excel(
        self, data: dict, filename: str
    ) -> tuple[bytes, str, str]:
        """Export as Excel workbook."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Add summary data
            ws_summary["A1"] = "AlphaGenome Analysis Report"
            ws_summary["A1"].font = Font(bold=True, size=14)
            ws_summary["A3"] = "Generated:"
            ws_summary["B3"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

            variant = data.get("request_params", {}).get("variant", "N/A")
            ws_summary["A4"] = "Variant:"
            ws_summary["B4"] = variant

            # Summary section
            summary = data.get("summary", {})
            if isinstance(summary, dict):
                row = 6
                ws_summary[f"A{row}"] = "Metric"
                ws_summary[f"B{row}"] = "Value"
                ws_summary[f"A{row}"].font = header_font
                ws_summary[f"A{row}"].fill = header_fill
                ws_summary[f"B{row}"].font = header_font
                ws_summary[f"B{row}"].fill = header_fill

                row += 1
                for key, value in summary.items():
                    if isinstance(value, list):
                        value = ", ".join(str(v) for v in value)
                    ws_summary[f"A{row}"] = key.replace("_", " ").title()
                    ws_summary[f"B{row}"] = str(value)
                    row += 1

            # Scores sheet
            scores = data.get("scores", [])
            if scores:
                ws_scores = wb.create_sheet("Gene Scores")

                headers = ["Gene Name", "Gene ID", "Strand", "Tissue", "Raw Score", "Quantile", "Interpretation"]
                for col, header in enumerate(headers, 1):
                    cell = ws_scores.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                for row_idx, score in enumerate(scores, 2):
                    if isinstance(score, dict):
                        ws_scores.cell(row=row_idx, column=1, value=score.get("gene_name", ""))
                        ws_scores.cell(row=row_idx, column=2, value=score.get("gene_id", ""))
                        ws_scores.cell(row=row_idx, column=3, value=score.get("strand", ""))
                        ws_scores.cell(row=row_idx, column=4, value=score.get("tissue", ""))
                        ws_scores.cell(row=row_idx, column=5, value=score.get("raw_score", 0))
                        ws_scores.cell(row=row_idx, column=6, value=score.get("quantile_score", 0))
                        ws_scores.cell(row=row_idx, column=7, value=score.get("interpretation", ""))

                # Auto-width columns
                for col in ws_scores.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    ws_scores.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return (
                buffer.read(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"{filename}.xlsx",
            )

        except ImportError:
            # Fallback to CSV
            return self._export_csv(data, filename)

    def get_copyable_text(self, data: dict, format: str = "markdown") -> str:
        """
        Get text that can be easily copied to clipboard.

        Args:
            data: The data to format
            format: 'markdown', 'json', or 'tsv'

        Returns:
            Formatted string ready for copy-paste
        """
        if format == "markdown":
            content, _, _ = self._export_markdown(data, "clipboard")
            return content.decode("utf-8")
        elif format == "json":
            return json.dumps(data, indent=2, default=str)
        elif format == "tsv":
            content, _, _ = self._export_tsv(data, "clipboard")
            return content.decode("utf-8")
        else:
            return json.dumps(data, indent=2, default=str)


# Singleton instance
export_service = ExportService()
